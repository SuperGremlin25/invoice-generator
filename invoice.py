# invoice-flow: Professional Invoice Generator
# White-labeled and maintained by Digital Insurgent Media
# Based on original work by Khaled Mahmoud (k5602)
# Licensed under the MIT License
# Version 2.0 - Premium Edition

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from tkcalendar import DateEntry
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.units import inch
import pandas as pd
import os
import json
import smtplib
import ssl
from email import encoders
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.application import MIMEApplication
from datetime import datetime
from PIL import Image as PILImage

# ---------------------------- Constants & Globals ----------------------------
LANGUAGE = "English"
items = []  # List to hold invoice items
theme = "light"
DEFAULT_CONFIG = {
    "company_name": "",
    "address": "",
    "city": "",
    "country": "",
    "phone": "",
    "email": "",
    "website": "",
    "tax_id": "",
    "currency": "USD",
    "logo_path": ""
}

# ---------------------------- Helper Functions ----------------------------
def load_config():
    try:
        if os.path.exists("config.json"):
            with open("config.json", "r") as f:
                loaded_config = json.load(f)
                # Ensure all default keys exist
                return {**DEFAULT_CONFIG, **loaded_config}
    except Exception as e:
        print(f"Error loading config: {e}")
    return DEFAULT_CONFIG.copy()

def save_config():
    try:
        # Update the global config dictionary
        config.update({
            "company_name": company_name_var.get(),
            "address": address_var.get(),
            "city": city_var.get(),
            "country": country_var.get(),
            "phone": phone_var.get(),
            "email": email_var.get(),
            "website": website_var.get(),
            "tax_id": tax_id_var.get(),
            "currency": currency_var.get() or "USD",  # Ensure USD is used if not set
            "logo_path": config.get("logo_path", "")
        })
        
        with open("config.json", "w") as f:
            json.dump(config, f, indent=4)
        
        # Update UI elements that depend on config
        update_items_table()
        
        messagebox.showinfo("Success", "Company settings saved successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save settings: {e}")

# ---------------------------- PDF Generation ----------------------------
def export_to_excel():
    """Export invoice data to Excel/CSV"""
    if not items:
        messagebox.showwarning("No Data", "No invoice items to export!")
        return
    
    try:
        # Get file path to save
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("CSV files", "*.csv"),
                ("All files", "*.*")
            ],
            title="Save Invoice Data As"
        )
        
        if not file_path:  # User cancelled
            return
            
        # Prepare data
        data = {
            "Item": [item[1] for item in items],
            "Quantity": [float(item[2]) for item in items],
            "Unit Price": [float(item[3]) for item in items],
            "Total": [float(item[2]) * float(item[3]) for item in items]
        }
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Calculate summary
        subtotal = df["Total"].sum()
        tax_rate = float(entry_tax.get() or 0)
        tax = subtotal * (tax_rate / 100)
        total = subtotal + tax
        
        # Add summary rows
        summary = pd.DataFrame({
            "Item": ["", "Subtotal", f"Tax ({tax_rate}%)", "Total"],
            "Quantity": ["", "", "", ""],
            "Unit Price": ["", "", "", ""],
            "Total": ["", subtotal, tax, total]
        })
        
        df = pd.concat([df, summary], ignore_index=True)
        
        # Export based on file extension
        if file_path.endswith('.csv'):
            df.to_csv(file_path, index=False, float_format='%.2f')
            messagebox.showinfo("Success", f"Invoice data exported to:\n{file_path}")
        else:  # Default to Excel
            try:
                # Use openpyxl for Excel export
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Invoice')
                    
                    # Get the workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Invoice']
                    
                    # Import styling modules
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    # Create styles
                    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFF')
                    bold_font = Font(bold=True)
                    
                    # Format headers
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(1, col_num)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Format summary rows (last 3 rows)
                    for row in range(len(df) - 2, len(df) + 1):
                        for col in range(1, len(df.columns) + 1):
                            cell = worksheet.cell(row + 1, col)  # +1 because Excel is 1-indexed
                            cell.font = bold_font
                    
                    # Set column widths
                    for idx, col in enumerate(df.columns):
                        # Get maximum length of any item in the column
                        max_length = max(df[col].astype(str).apply(len).max(), len(str(col)))
                        # Set column width based on max length
                        column_letter = worksheet.cell(1, idx + 1).column_letter
                        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 30)
                    
                    # Format currency columns (Unit Price and Total)
                    for row in range(2, len(df) + 2):  # Excel is 1-indexed, and header is row 1
                        # Format Unit Price column
                        unit_price_col = worksheet.cell(1, 3).column_letter
                        worksheet[f'{unit_price_col}{row}'].number_format = '#,##0.00'
                        
                        # Format Total column
                        total_col = worksheet.cell(1, 4).column_letter
                        worksheet[f'{total_col}{row}'].number_format = '#,##0.00'
                
                messagebox.showinfo("Success", f"Invoice data exported to:\n{file_path}")
            except Exception as excel_error:
                messagebox.showerror("Excel Export Error", f"Failed to export to Excel: {str(excel_error)}\n\nTrying CSV export as fallback...")
                
                # Fallback to CSV if Excel export fails
                csv_path = file_path.rsplit(".", 1)[0] + ".csv"
                df.to_csv(csv_path, index=False, float_format='%.2f')
                messagebox.showinfo("Success", f"Invoice data exported to CSV as fallback:\n{csv_path}")
        
    except Exception as e:
        messagebox.showerror("Export Error", f"Failed to export data: {str(e)}")


def generate_pdf(filename=None, show_dialog=True):
    """Generate a PDF invoice with proper formatting and error handling"""
    try:
        # Get customer and invoice details
        customer_name = entry_customer.get().strip()
        invoice_number = entry_invoice_number.get().strip()
        invoice_date = date_entry.get()
        tax_rate = float(entry_tax.get() or 0)
        
        # Company details (from global config)
        company_name = company_name_var.get() or "Your Company"
        company_address = address_var.get() or ""
        company_city = city_var.get() or ""
        company_country = country_var.get() or ""
        company_phone = phone_var.get() or ""
        company_email = email_var.get() or ""
        company_website = website_var.get() or ""
        company_tax_id = tax_id_var.get() or ""
        
        # Currency - ensure it defaults to USD
        currency = currency_var.get() or "USD"
        
        if not items:
            messagebox.showwarning("No Items", "Please add at least one item to the invoice.")
            return False
        
        # If filename is not provided and show_dialog is True, ask for a filename
        if not filename and show_dialog:
            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                title="Save PDF Invoice As"
            )
            
            if not filename:  # User cancelled
                return False
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(os.path.abspath(filename)), exist_ok=True)
        
        # Create a PDF document
        doc = SimpleDocTemplate(filename, pagesize=letter)
        
        # Container for the PDF elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = styles["Title"]
        heading_style = styles["Heading2"]
        normal_style = styles["Normal"]
        italic_style = styles["Italic"]
        
        # Create custom styles
        right_style = ParagraphStyle(
            name='RightAlign',
            parent=styles['Normal'],
            alignment=TA_RIGHT  # Using reportlab constants for alignment
        )
        
        # Add company logo if available
        logo_path = config.get("logo_path", "")
        if logo_path and os.path.exists(logo_path):
            try:
                img = PILImage.open(logo_path)
                width, height = img.size
                aspect = height / width
                img_width = 1.5 * inch
                img_height = img_width * aspect
                logo = Image(logo_path, width=img_width, height=img_height)
                elements.append(logo)
                elements.append(Spacer(1, 12))
            except Exception as e:
                print(f"Error loading logo: {e}")
        
        # Add title
        elements.append(Paragraph(f"{company_name}", title_style))
        elements.append(Spacer(1, 12))
        
        # Company details
        if company_address or company_city or company_country:
            address_str = ", ".join(filter(None, [company_address, company_city, company_country]))
            elements.append(Paragraph(f"<b>Address:</b> {address_str}", normal_style))
        
        if company_phone:
            elements.append(Paragraph(f"<b>Phone:</b> {company_phone}", normal_style))
        
        if company_email:
            elements.append(Paragraph(f"<b>Email:</b> {company_email}", normal_style))
        
        if company_website:
            elements.append(Paragraph(f"<b>Website:</b> {company_website}", normal_style))
        
        if company_tax_id:
            elements.append(Paragraph(f"<b>Tax ID:</b> {company_tax_id}", normal_style))
        
        elements.append(Spacer(1, 24))
        
        # Invoice header
        elements.append(Paragraph("INVOICE", heading_style))
        elements.append(Spacer(1, 12))
        
        # Invoice details in a table format
        invoice_data = [
            [Paragraph("<b>Invoice #:</b>", normal_style), invoice_number],
            [Paragraph("<b>Date:</b>", normal_style), invoice_date],
            [Paragraph("<b>Customer:</b>", normal_style), customer_name]
        ]
        
        invoice_table = Table(invoice_data, colWidths=[100, 150])
        invoice_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ]))
        
        elements.append(invoice_table)
        elements.append(Spacer(1, 24))
        
        # Invoice items table
        items_data = [["Item", "Quantity", f"Unit Price ({currency})", f"Total ({currency})"]]
        
        subtotal = 0
        for item_data in items:
            name = item_data[1]
            quantity = float(item_data[2])
            price = float(item_data[3])
            total = quantity * price
            subtotal += total
            
            items_data.append([name, f"{quantity:.2f}", f"{price:.2f}", f"{total:.2f}"])
        
        # Calculate tax and total
        tax_amount = subtotal * (tax_rate / 100)
        total_amount = subtotal + tax_amount
        
        # Add summary rows
        items_data.append(["", "", "Subtotal:", f"{subtotal:.2f}"])
        items_data.append(["", "", f"Tax ({tax_rate}%):", f"{tax_amount:.2f}"])
        items_data.append(["", "", "Total:", f"{total_amount:.2f}"])
        
        # Create the table with the items
        col_widths = [doc.width * 0.4, doc.width * 0.15, doc.width * 0.2, doc.width * 0.25]
        items_table = Table(items_data, colWidths=col_widths)
        
        # Apply styles to the table
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -4), colors.beige),
            ('GRID', (0, 0), (-1, -4), 1, colors.black),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, -3), (-1, -1), 'Helvetica-Bold'),
        ]
        items_table.setStyle(TableStyle(table_style))
        
        elements.append(items_table)
        elements.append(Spacer(1, 48))
        
        # Footer with terms
        elements.append(Paragraph("<b>Terms & Conditions</b>", normal_style))
        elements.append(Paragraph("Payment is due within 30 days.", normal_style))
        elements.append(Paragraph("Please make checks payable to the company name above.", normal_style))
        
        # Build the PDF
        doc.build(elements)
        
        if show_dialog:
            messagebox.showinfo("Success", f"Invoice PDF generated successfully at:\n{filename}")
            
            # Try to open the PDF automatically
            try:
                import subprocess
                import platform
                
                if platform.system() == 'Windows':
                    os.startfile(os.path.normpath(filename))
                elif platform.system() == 'Darwin':  # macOS
                    subprocess.run(['open', filename])
                else:  # Linux
                    subprocess.run(['xdg-open', filename])
            except Exception as e:
                print(f"Failed to open PDF: {e}")
                messagebox.showinfo("PDF Created", f"PDF created but couldn't be opened automatically.\nLocation: {filename}")
        
        return True
        
    except Exception as e:
        if show_dialog:
            messagebox.showerror("PDF Generation Error", f"Failed to generate PDF: {str(e)}")
        print(f"PDF generation error: {e}")
        return False

# ---------------------------- GUI Functions ----------------------------
def add_item():
    item_name = entry_item.get()
    quantity = entry_quantity.get()
    price = entry_price.get()
    
    if item_name and quantity and price:
        try:
            # Add item with an index number
            item_index = len(items) + 1
            items.append((item_index, item_name, float(quantity), float(price)))
            update_items_table()
            entry_item.delete(0, tk.END)
            entry_quantity.delete(0, tk.END)
            entry_price.delete(0, tk.END)
        except ValueError:
            messagebox.showerror("Error", "Quantity and Price must be numbers!")
    else:
        messagebox.showerror("Error", "All item fields are required!")

def update_items_table():
    # Update currency label - ensure we use USD as default if not set
    currency = config.get("currency", "USD")
    if price_label:
        price_label.config(text=f"Price ({currency}):")
        
    # Update table header
    items_table.heading("Price", text=f"Price ({currency})")
    
    # Clear table
    for item in items_table.get_children():
        items_table.delete(item)
        
    # Populate table with items
    subtotal = 0.0
    for item_data in items:
        # Unpack tuple which now includes the index
        item_index, item_name, quantity, price = item_data
        item_total = quantity * price
        subtotal += item_total
        items_table.insert("", "end", values=(item_index, item_name, quantity, f"{price:.2f}"))
    
    # Calculate and display total
    tax_rate = 0.0
    try:
        tax_rate = float(entry_tax.get()) / 100.0
    except (ValueError, AttributeError):
        pass
    
    tax_amount = subtotal * tax_rate
    total = subtotal + tax_amount
    
    # Update summary display
    if "total_frame" not in globals():
        # Create the totals display frame if it doesn't exist
        global total_frame, subtotal_label, tax_label, total_label
        total_frame = ttk.LabelFrame(invoice_tab, text="Invoice Summary", padding=10)
        total_frame.pack(side="bottom", fill="x", padx=10, pady=5)
        
        # Create grid layout for the summary
        total_frame.columnconfigure(1, weight=1)
        
        # Subtotal row
        ttk.Label(total_frame, text="Subtotal:").grid(row=0, column=0, sticky="w", padx=5)
        subtotal_label = ttk.Label(total_frame, text="")
        subtotal_label.grid(row=0, column=1, sticky="e", padx=5)
        
        # Tax row
        ttk.Label(total_frame, text="Tax:").grid(row=1, column=0, sticky="w", padx=5)
        tax_label = ttk.Label(total_frame, text="")
        tax_label.grid(row=1, column=1, sticky="e", padx=5)
        
        # Total row
        ttk.Label(total_frame, text="Total:", font=("Segoe UI", 10, "bold")).grid(row=2, column=0, sticky="w", padx=5)
        total_label = ttk.Label(total_frame, text="", font=("Segoe UI", 10, "bold"))
        total_label.grid(row=2, column=1, sticky="e", padx=5)
    
    # Update the labels with the calculated values
    subtotal_label.config(text=f"{currency} {subtotal:.2f}")
    tax_label.config(text=f"{currency} {tax_amount:.2f}")
    total_label.config(text=f"{currency} {total:.2f}")


def load_logo(image_path, size=(150, 80)):
    """Load and resize logo image"""
    if not os.path.exists(image_path):
        return None
    try:
        # Use PIL to open and resize the image
        from PIL import Image, ImageTk
        image = Image.open(image_path)
        image = image.resize(size, Image.LANCZOS)
        photo = ImageTk.PhotoImage(image)
        return photo
    except Exception as e:
        print(f"Error loading logo: {e}")
        return None

def select_logo():
    """Allow user to select a logo file"""
    file_path = filedialog.askopenfilename(
        title="Select Logo",
        filetypes=[
            ("Image files", "*.png *.jpg *.jpeg *.gif *.bmp"),
            ("All files", "*.*")
        ]
    )
    
    if file_path:
        # Update the config with the new logo path
        config["logo_path"] = file_path
        save_config()
        
        # Display the logo
        company_logo = load_logo(file_path)
        if company_logo:
            logo_label.config(image=company_logo)
            logo_label.image = company_logo  # Keep a reference
        else:
            logo_label.config(text="Failed to load logo")

def remove_item():
    """Remove the last item from the items list"""
    if items:
        items.pop()
        update_items_table()

# Create main window and notebook tabs
root = tk.Tk()
root.title("Invoice Generator - Premium Edition")
root.geometry("1000x750")

# Set theme colors for a more premium look
style = ttk.Style()
style.theme_use('clam')  # Use the clam theme as a base

# Define premium theme colors
primary_color = "#2c3e50"  # Dark blue
secondary_color = "#3498db"  # Light blue
accent_color = "#e74c3c"  # Red accent
text_color = "#2c3e50"  # Dark text
light_text = "#ecf0f1"  # Light text for dark backgrounds

# Configure styles for different widgets
style.configure('TLabel', foreground=text_color, font=('Segoe UI', 10))
style.configure('TButton', font=('Segoe UI', 10, 'bold'))
style.configure('Accent.TButton', foreground=light_text, background=accent_color)
style.configure('TNotebook', background=primary_color)
style.configure('TNotebook.Tab', background=primary_color, foreground=light_text, padding=[10, 5], font=('Segoe UI', 10))
style.configure('TFrame', background='#ffffff')

# Load config first
config = load_config()

# Create notebook (tabbed interface)
nb = ttk.Notebook(root)
nb.pack(fill="both", expand=True, padx=10, pady=10)

# Create tabs
invoice_tab = ttk.Frame(nb)
settings_tab = ttk.Frame(nb)
reports_tab = ttk.Frame(nb)  # Premium feature tab
contacts_tab = ttk.Frame(nb)  # Premium feature tab

# Add tabs to notebook
nb.add(invoice_tab, text="Invoice")
nb.add(settings_tab, text="Settings")
nb.add(reports_tab, text="Reports ⭐")
nb.add(contacts_tab, text="Contacts ⭐")

# Add premium feature notices to premium tabs
ttk.Label(
    reports_tab, 
    text="Reports Feature - Premium Edition", 
    font=("Segoe UI", 14, "bold")
).pack(pady=50)

ttk.Label(
    reports_tab, 
    text="Track your sales performance, view customer history,\nand generate financial reports.", 
    font=("Segoe UI", 12)
).pack()

ttk.Label(
    contacts_tab, 
    text="Customer Management - Premium Edition", 
    font=("Segoe UI", 14, "bold")
).pack(pady=50)

ttk.Label(
    contacts_tab, 
    text="Manage your customer database, track communication history,\nand set follow-up reminders.", 
    font=("Segoe UI", 12)
).pack()

# Invoice tab layout - split into left and right frames
frame_left = ttk.LabelFrame(invoice_tab, text="Invoice Details", padding=10)
frame_left.pack(side="left", fill="both", expand=True, padx=5, pady=5)

# Invoice details entry fields
ttk.Label(frame_left, text="Customer:").grid(row=0, column=0, sticky="w")
entry_customer = ttk.Entry(frame_left, width=30)
entry_customer.grid(row=0, column=1, pady=5)

ttk.Label(frame_left, text="Invoice Number:").grid(row=1, column=0, sticky="w")
entry_invoice_number = ttk.Entry(frame_left, width=30)
entry_invoice_number.grid(row=1, column=1, pady=5)

ttk.Label(frame_left, text="Date:").grid(row=2, column=0, sticky="w")
date_entry = DateEntry(frame_left, width=29, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')
date_entry.grid(row=2, column=1, pady=5)

ttk.Label(frame_left, text="Tax Rate (%):").grid(row=3, column=0, sticky="w")
entry_tax = ttk.Entry(frame_left, width=30)
entry_tax.grid(row=3, column=1, pady=5)
entry_tax.insert(0, "0")  # Default to 0%

# Right frame for item entry
frame_right = ttk.LabelFrame(invoice_tab, text="Add Items", padding=10)
frame_right.pack(side="right", fill="y", padx=5, pady=5)

# Item entry fields
ttk.Label(frame_right, text="Item:").grid(row=0, column=0, sticky="w")
entry_item = ttk.Entry(frame_right)
entry_item.grid(row=0, column=1, pady=5)

ttk.Label(frame_right, text="Quantity:").grid(row=1, column=0, sticky="w")
entry_quantity = ttk.Entry(frame_right)
entry_quantity.grid(row=1, column=1, pady=5)

price_label = ttk.Label(frame_right, text=f"Price ({config.get('currency', 'USD')}):")
price_label.grid(row=2, column=0, sticky="w")
entry_price = ttk.Entry(frame_right)
entry_price.grid(row=2, column=1, pady=5)

ttk.Button(frame_right, text="Add Item", command=add_item).grid(row=3, column=0, pady=10)
ttk.Button(frame_right, text="Remove Last Item", command=remove_item).grid(row=3, column=1, pady=10)

columns = ("#", "Item", "Quantity", "Price")
items_table = ttk.Treeview(frame_left, columns=columns, show="headings", height=10)
items_table.heading("#", text="#")
items_table.heading("Item", text="Item")
items_table.heading("Quantity", text="Quantity")
items_table.heading("Price", text=f"Price ({config.get('currency', 'USD')})")
items_table.grid(row=4, column=0, columnspan=2, pady=10)

def send_email():
    # First generate the PDF
    pdf_filename = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf")],
        initialfile=f"{entry_invoice_number.get().replace(' ', '_')}.pdf"
    )
    
    if not pdf_filename:  # User cancelled file dialog
        return False 
    
    # Generate the PDF file for email attachment
    if not generate_pdf(filename=pdf_filename, show_dialog=False):
        messagebox.showerror("Error", "Failed to generate PDF for email attachment!")
        return False
        
    # Verify the PDF was created
    if not os.path.exists(pdf_filename):
        messagebox.showerror("Error", "PDF file was not created successfully!")
        return False
    
    # Get recipient email
    recipient = simpledialog.askstring("Email", "Enter recipient email:")
    if not recipient:
        return False  # User cancelled dialog
    
    # Get sender email and password
    sender_email = simpledialog.askstring("Email", "Enter your Gmail address:")
    if not sender_email:
        return False
    
    password = simpledialog.askstring("Email", "Enter your Gmail app password:", show="*")
    if not password:
        return False
    
    # Setup email message
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient
    msg['Subject'] = f"Invoice {entry_invoice_number.get()}"
    
    # Email body
    body = f"Please find attached invoice {entry_invoice_number.get()} for {entry_customer.get()}.\n\nThank you,\n{config.get('company_name', '')}"
    msg.attach(MIMEText(body, 'plain'))
    
    # Attach PDF
    with open(pdf_filename, 'rb') as attachment:
        part = MIMEBase('application', 'pdf')
        part.set_payload(attachment.read())
    
        encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition',
            f"attachment; filename= {os.path.basename(pdf_filename)}",
        )
        msg.attach(part)
        
        # Send email
        try:
            # Create a secure SSL/TLS connection to Gmail SMTP server
            server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
            server.ehlo()
            
            # Login to your Gmail account
            server.login(sender_email, password)
            
            # Send the email
            text = msg.as_string()
            server.sendmail(sender_email, recipient, text)
            server.quit()
            messagebox.showinfo("Success", "Email sent successfully!")
            return True
        except smtplib.SMTPAuthenticationError:
            messagebox.showerror("Email Error", "Authentication failed. Please check your email and app password.\n\nNote: You need to use an App Password if you have 2FA enabled on your Google account.")
            return False
        except Exception as e:
            messagebox.showerror("Email Error", f"Failed to send email: {e}")
            return False

# Create a frame for totals display that will be updated by update_items_table()
# Note: The actual frame will be created in update_items_table if needed

# Bottom buttons in invoice tab with premium styling
button_frame = ttk.Frame(invoice_tab)
button_frame.pack(fill="x", padx=10, pady=15)

# Left side buttons for main actions
left_btn_frame = ttk.Frame(button_frame)
left_btn_frame.pack(side="left")

# Create styled buttons
generate_btn = ttk.Button(left_btn_frame, text="Generate PDF", command=generate_pdf, style="Accent.TButton")
generate_btn.pack(side="left", padx=5)

export_btn = ttk.Button(left_btn_frame, text="Export to Excel/CSV", command=export_to_excel)
export_btn.pack(side="left", padx=5)

clear_btn = ttk.Button(left_btn_frame, text="Clear All Items", command=lambda: [items.clear(), update_items_table()])
clear_btn.pack(side="left", padx=5)

# Right side button for email
right_btn_frame = ttk.Frame(button_frame)
right_btn_frame.pack(side="right")

email_btn = ttk.Button(right_btn_frame, text="📧 Send Invoice", command=send_email)
email_btn.pack(side="right", padx=5)

# Premium feature indicator
premium_label = ttk.Label(button_frame, text="Premium Features Available", foreground="#e67e22", font=("Segoe UI", 9, "italic"))
premium_label.pack(side="right", padx=20)


# Create variables (config already loaded above)
# These variables will be used for the settings tab

# Create variables
company_name_var = tk.StringVar(value=config["company_name"])
address_var = tk.StringVar(value=config["address"])
city_var = tk.StringVar(value=config["city"])
country_var = tk.StringVar(value=config["country"])
phone_var = tk.StringVar(value=config["phone"])
email_var = tk.StringVar(value=config["email"])
website_var = tk.StringVar(value=config["website"])
tax_id_var = tk.StringVar(value=config["tax_id"])
currency_var = tk.StringVar(value=config["currency"])

# Logo frame
logo_frame = ttk.Frame(settings_tab, padding=10)
logo_frame.pack(fill="x", padx=10, pady=5)

# Logo display
logo_label = ttk.Label(logo_frame, text="No Logo Selected")
logo_label.pack(side="left", padx=10)

# Load existing logo if exists
if os.path.exists(config.get("logo_path", "")):
    company_logo = load_logo(config["logo_path"])
    if company_logo:
        logo_label.config(image=company_logo)
        logo_label.image = company_logo

# Logo upload button
ttk.Button(
    logo_frame, 
    text="Upload Logo", 
    command=select_logo
).pack(side="left", padx=10)

# Form frame
form_frame = ttk.Frame(settings_tab, padding=20)
form_frame.pack(fill="both", expand=True)

# Company Information
ttk.Label(form_frame, text="Company Information", font=("Arial", 12, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 20), sticky="w")

# Left column
ttk.Label(form_frame, text="Company Name:").grid(row=1, column=0, sticky="w", pady=5)
ttk.Entry(form_frame, textvariable=company_name_var, width=40).grid(row=1, column=1, sticky="w", pady=5)

ttk.Label(form_frame, text="Address:").grid(row=2, column=0, sticky="w", pady=5)
ttk.Entry(form_frame, textvariable=address_var, width=40).grid(row=2, column=1, sticky="w", pady=5)

ttk.Label(form_frame, text="City:").grid(row=3, column=0, sticky="w", pady=5)
ttk.Entry(form_frame, textvariable=city_var, width=40).grid(row=3, column=1, sticky="w", pady=5)

ttk.Label(form_frame, text="Country:").grid(row=4, column=0, sticky="w", pady=5)
ttk.Entry(form_frame, textvariable=country_var, width=40).grid(row=4, column=1, sticky="w", pady=5)

# Right column
ttk.Label(form_frame, text="Phone:").grid(row=1, column=2, sticky="w", pady=5, padx=(30, 0))
ttk.Entry(form_frame, textvariable=phone_var, width=30).grid(row=1, column=3, sticky="w", pady=5)

ttk.Label(form_frame, text="Email:").grid(row=2, column=2, sticky="w", pady=5, padx=(30, 0))
ttk.Entry(form_frame, textvariable=email_var, width=30).grid(row=2, column=3, sticky="w", pady=5)

ttk.Label(form_frame, text="Website:").grid(row=3, column=2, sticky="w", pady=5, padx=(30, 0))
ttk.Entry(form_frame, textvariable=website_var, width=30).grid(row=3, column=3, sticky="w", pady=5)

ttk.Label(form_frame, text="Tax ID:").grid(row=4, column=2, sticky="w", pady=5, padx=(30, 0))
ttk.Entry(form_frame, textvariable=tax_id_var, width=30).grid(row=4, column=3, sticky="w", pady=5)

ttk.Label(form_frame, text="Currency:").grid(row=5, column=2, sticky="w", pady=5, padx=(30, 0))
currency_combo = ttk.Combobox(form_frame, textvariable=currency_var, values=["USD", "EGP", "EUR", "GBP"], width=27, state="readonly")
currency_combo.grid(row=5, column=3, sticky="w", pady=5)

# Save button
button_frame = ttk.Frame(settings_tab)
button_frame.pack(fill="x", padx=5, pady=10)

ttk.Button(button_frame, text="Save Settings", command=save_config).pack(side="left", padx=5)
def reset_defaults():
    for var in [company_name_var, address_var, city_var, country_var, 
               phone_var, email_var, website_var, tax_id_var]:
        var.set("")
    currency_var.set("USD")

ttk.Button(button_frame, text="Reset to Defaults", 
          command=reset_defaults).pack(side="left", padx=5)


# Ensure USD is the first option and default currency
currency_combo['values'] = ["USD", "EGP", "EUR", "GBP"]
    
# Explicitly set USD as the selected currency
if currency_var.get() != "USD":
    currency_var.set("USD")
    config["currency"] = "USD"
    save_config()

# Initial call to set the currency on startup (after all UI elements are created)
root.update()  # Ensure all UI elements are fully created
update_items_table()

root.mainloop()